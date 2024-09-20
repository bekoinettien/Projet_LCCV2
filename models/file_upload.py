# from odoo import models, fields, api, _
# import base64
# import os
# import sys
# sys.path.append('/usr/lib/python3/dist-packages/demo')
# sys.path.append('D:\Odoo17\python\Lib\site-packages')
# import msoffcrypto
# #from io import BytesIO
# import openpyxl
# from odoo.exceptions import UserError
#
# class FileUpload(models.Model):
#     _name = 'file.upload'
#     _description = 'File Upload'
#
#     file_data = fields.Binary(string='File', required=True)
#
#     @api.model
#     def create(self, vals):
#         record = super(FileUpload, self).create(vals)
#         if record.file_data:
#             file_path = self.save_file_to_disk(record.file_data)
#             self.import_pesse(file_path)
#         return record
#
#     def save_file_to_disk(self, file_data):
#         decoded_file = base64.b64decode(file_data)
#         fixed_file_name = "lcctemporyfile.xlsx"
#         file_path = os.path.join('/cryptage', fixed_file_name)
#
#         with open(file_path, 'wb') as file:
#             file.write(decoded_file)
#
#         fichier_crypter = open(r"/cryptage/lcctemporyfile.xlsx", "rb")
#         file = msoffcrypto.OfficeFile(fichier_crypter)
#         file.load_key(password="IPACRCI")
#
#         # decrypted_file_path = os.path.join('/cryptage', "decrypter.xlsx")
#         with open(r"/cryptage/decrypter.xlsx", 'wb') as f:
#             file.decrypt(f)
#
#         return r"/cryptage/decrypter.xlsx"
#
#     def import_pesse(self, path):
#         try:
#             wb = openpyxl.load_workbook(filename=r"/cryptage/decrypter.xlsx", read_only=True)
#             ws = wb.active
#
#             for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,
#                                        max_col=None, values_only=True):
#
#                 search = self.env['weight.weight'].search([('name', '=', record[0])])
#
#                 if not search:
#                     self.env['weight.weight'].create({
#                         'name': record[0],
#                         'date': record[1],
#                         'code_farmer': record[2],
#                         'qty': record[5],
#                     })
#                 else:
#                     raise UserError(_('Pesée déjà importée: %s' % record[1]))
#         except Exception as e:
#             raise UserError(_('Please insert a valid file. Error: %s' % str(e)))
#         return
from odoo import models, fields, api, _
import base64
import os
import sys
# sys.path.append('/usr/lib/python3/dist-packages/demo')
sys.path.append('D:\Odoo17\python\Lib\site-packages')
import msoffcrypto
import openpyxl
from odoo.exceptions import UserError
# import logging
#
# _logger = logging.getLogger(__name__)

class FileUpload(models.Model):
    _name = 'file.upload'
    _description = 'File Upload'


    file_data = fields.Binary(string='File', required=True)


    @api.model
    def create(self, vals):
        record = super(FileUpload, self).create(vals)
        if record.file_data:
            self.save_file_to_disk(record.file_data)
        return record

    def save_file_to_disk(self, file_data):
        decoded_file = base64.b64decode(file_data)
        fixed_file_name = "lcctemporyfile.xlsx"
        #file_path = os.path.join('/cryptage', fixed_file_name)  # Change the path to your desired directory
        file_path = os.path.join('D:\BTS2', fixed_file_name)  # Change the path to your desired directory

        with open(file_path, 'wb') as file:
            file.write(decoded_file)

        # fichier_crypter = open(r"/cryptage/lcctemporyfile.xlsx", "rb")
        fichier_crypter = open(r"D:\BTS2\lcctemporyfile.xlsx", "rb")
        file = msoffcrypto.OfficeFile(fichier_crypter)
        file.load_key(password="IPACRCI")
        #decrypted_file_path = r"/cryptage/decrypter.xlsx"
        decrypted_file_path = r"D:\BTS2\decrypter.xlsx"
        with open(decrypted_file_path, "wb") as f:
         file.decrypt(f)
         self.load_data_to_model(decrypted_file_path)
        return fixed_file_name

    def load_data_to_model(self, decrypted_file_path):
        workbook = openpyxl.load_workbook(decrypted_file_path)
        sheet = workbook.active

        # Assurez-vous que les colonnes sont bien ordonnées comme attendu
        for row in sheet.iter_rows(min_row=2, values_only=True):  # min_row=2 pour ignorer l'en-tête
            # Adaptez ces valeurs selon les colonnes de votre fichier Excel
            code = row[1]
            datelive = row[2]
            codeplat = row[3]
            quantite = row[6]
            # code = row[12]
            # datelive = row[2]
            # codeplat = row[3]
            # quantite = row[6]


            # Ajoutez d'autres champs nécessaires

            # Créer ou mettre à jour l'enregistrement dans le modèle "weight.weight"
            search = self.env['weight.weight'].search([('name', '=', code)], limit=1)
            if not search:
                # Créer l'enregistrement seulement s'il n'existe pas
                self.env['weight.weight'].create({
                        'name': code,
                        'date': datelive,
                        'code_farmer': codeplat,
                        'qty': quantite,

                        # Remplissez les autres champs ici
                    })
            else:
                 raise UserError(_('Pesée déjà importée: %s' % row[1]))
            # except Exception as e:
            #            raise UserError(_('Please insert a valid file. Error: %s' % str(e)))
